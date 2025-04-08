import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Load JSON data ===
with open("../jsons/echantillon.json", "r", encoding="utf-8") as file:
    videos = json.load(file)

with open("../jsons/channel_names.json", "r", encoding="utf-8") as file:
    channel_names = json.load(file)

with open("../jsons/categories.json", "r", encoding="utf-8") as file:
    categorie_labels = json.load(file)

# === Prepare Data ===
video_data = []
for video in videos:
    channel_name = channel_names.get(video["id_chaine"], "Unknown Channel")
    category_name = categorie_labels.get(str(video["youtubeCategorie"]), "Unknown Category")
    video_data.append({
        "Video": video["titre_video"],
        "Lien": f"https://www.youtube.com/watch?v={video['id_video']}",
        "Chaine (Nom)": channel_name,
        "Categorie (Nom)": category_name,
        "CategorieId": video["youtubeCategorie"]
    })

df = pd.DataFrame(video_data)
df = df.sort_values(by="CategorieId", ascending=False)

# === Write to Excel first ===
excel_path = "../xls/videos_categories.xlsx"
df.to_excel(excel_path, index=False)

# === Open with openpyxl for formatting ===
wb = load_workbook(excel_path)
ws = wb.active

# Assign a unique color per CategorieId
from itertools import cycle

# Color palette (light pastel tones)
colors = cycle([
    "FFDDC1", "FFABAB", "FFC3A0", "D5AAFF", "B5EAD7", "C7CEEA", "E2F0CB",
    "FFDAC1", "CBAACB", "FFFFD1", "FFB7B2", "B5EAD7", "FF9AA2", "E0BBE4"
])

# Get all unique categories and assign a color
categorie_to_color = {}
for i, cat_id in enumerate(sorted(df["CategorieId"].unique(), reverse=True)):
    categorie_to_color[str(cat_id)] = next(colors)

# Apply color row by row (start at row 2 since row 1 is headers)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    categorie_id = str(row[4].value)  # 5th column = CategorieId
    fill = PatternFill(start_color=categorie_to_color[categorie_id],
                       end_color=categorie_to_color[categorie_id],
                       fill_type="solid")
    for cell in row:
        cell.fill = fill

# Save the formatted file
wb.save(excel_path)

print("✅ Excel file with colored rows by CategorieId saved at:", excel_path)
